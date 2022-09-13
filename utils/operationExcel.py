# -*- coding: utf-8 -*-
"""
@Time ： 2022/9/12 16:50
@Auth ： linxiaoyi
@File ：operationExcel.py
@IDE ：PyCharm

"""
import openpyxl
import pathlib
import logging
#读取excel表数据
class operationExcel():
    def __init__(self,filename):
        # self.workbook = None
        self.sheet = None
        self.rows = 0
        self.r = 0
        self.filename=filename
        self.workbook =self.open_excel()

    #打开excel表
    def open_excel(self):
        if not pathlib.Path(self.filename).is_file():
            print("%s not exist!" % (self.filename))
            return
        self.workbook = openpyxl.load_workbook(self.filename)
        return self.workbook

    def get_sheets(self):
        self.sheets = self.workbook.sheetnames
        return self.sheets

    def set_sheet(self, name):
        self.sheet = self.workbook[name]
        self.rows = self.sheet.max_row
        self.r = 0
        return
    #获取sheet表
    def get_sheet(self,name):
        self.sheet=self.workbook[name]
        return self.sheet

    # 获取指定单元格的值
    def get_value(self,row,column):
        value=self.sheet.cell(row=row,column=column).value
        return value

    def get_excel_datas(self):#应按行取值 =====
       self.sheets = self.get_sheets()
       data=[]
       for name in self.sheets:
           sheet = self.get_sheet(name)
           columns=sheet.max_column
           rows=sheet.max_row
           # print("rows是",rows)
           # for column in range(1,columns+1):
           # title= sheet[1]
           row_value=[]
           title=[]
           for row in range(1,rows+1):
               row_values=sheet[row]
               if row == 1:
                   for i in row_values:
                       title.append(i.value)
                       # print("title:",i.value)
               else:
                   for i in row_values:
                       row_value.append(i.value)
                       # print("row_value:",i.value)
               # row_value= sheet.cell(row=row,column=column).value
               # print("title是", title)
               #
               data_dict=dict(zip(title, row_value))
               print("data_dict是", data_dict)
               data.append(data_dict)
       # print(data)
       # return data

   #获取所有sheet表的单元格的值,并组合为字典
    def get_allvalue(self):
        self.sheets=self.get_sheets()
        logging.info("即将执行{}sheet表".format(self.sheets))
        temp_dict={}
        temp_value=None
        temp_key=None
        for name in self.sheets:
            sheet=self.get_sheet(name)
            for value in sheet.values:
                # print(value)
                if type(value[0]) is not int:
                # if value[0]:
                    temp_key = value
                else:
                    temp_value=value
                    temp_dict = dict(zip(temp_key, temp_value))
                    # print("字典是{}".format(temp_dict))
                    yield temp_dict

if __name__ == '__main__':
    filepath='H:\\Program Files (x86)\\Python_Project\\Api_Pytest\\data\\b.xlsx'
    excel=operationExcel(filepath)
    # for i in excel.get_allvalue():
    #   print(i)
    print(excel.get_excel_datas())




